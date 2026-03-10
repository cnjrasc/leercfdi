from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from decimal import Decimal, InvalidOperation
from django.views.generic import DetailView
from django.contrib import messages
from datetime import date
from datetime import datetime
from .templates.utils import convertir_fecha_iso
from .templates.utils import cfdi_to_excel
import csv
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
#import chardet
from django.db.models import Sum
from django.template.loader import render_to_string
#from weasyprint import HTML
import pandas as pd
import zipfile
from io import BytesIO
from lxml import etree
from django.conf import settings
import os

from .models import Contribuyente
from .forms import ContribuyenteForm
#from .models import Cfdi
#from .forms import CfdiForm
from .forms import CFDIUploadForm
from .forms import CFDIZipForm
from .templates.utils import extract_cfdi_from_xml
# Create your views here.

def inicio(request):
    return render(request, 'paginas/inicio.html')

def home(request):
    return render(request, 'home.html')

def contribuyente(request):
    contribuyentes = Contribuyente.objects.all()
    return render(request, 'contribuyente/index.html', {'contribyentes': contribuyentes} )

def altaem(request):
    formulario = ContribuyenteForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('empresa')
    return render(request, 'contribuyente/altaem.html', {'formulario': formulario})

def editarem(request, id):
    contribuyente = Contribuyente.objects.get(id=id)
    formulario = ContribuyenteForm(request.POST or None, request.FILES or None, instance=contribuyente)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('contribuyente')
    return render(request, 'contribuyente/editarem.html', {'formulario': formulario})
def eliminarem(request, id):
    contribuyente = Contribuyente.objects.get(id=id)
    contribuyente.delete()
    return redirect('contribuyente')

#def cfdi(request):
    #cfdis = Cfdi.objects.all()
    #return render(request, 'cfdi/index.html', {'cfdis': cfdis} )

def altaem(request):
    formulario = CfdiForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('cfdi')
    return render(request, 'cfdi/altaem.html', {'formulario': formulario})

def editarem(request, id):
    cfdi = Cfdi.objects.get(id=id)
    formulario = CfdiForm(request.POST or None, request.FILES or None, instance=cfdi)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('cfdi')
    return render(request, 'cfdi/editarem.html', {'formulario': formulario})
def eliminarem(request, id):
    cfdi = Cfdi.objects.get(id=id)
    cfdi.delete()
    return redirect('cffdi')

def total_expense(request):
    total_subtotal = Expense.objects.aggregate(total_importe=Sum('subtotal'))['total_subtotal']
    total_iva = Expense.objects.aggregate(total_iva=Sum('iva'))['total_iva']
    total_total = Expense.objects.aggregate(total_total=Sum('total'))['total_total']
    
    
    context = {
        'total_subtotal': total_subtotal,
        'total_iva': total_iva,
        'total_total': total_total,

        'expense': Expense.objects.all()
    }
    return render(request, 'expense/index.html', context)

def guardar_expense(request):
    if request.method == 'POST':
        formulario = ExpenseForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('guardar_expense')
    else:
        formulario = ExpenseForm()
    return render(request, 'expense/guardar_expense.html', {'formulario': formulario})

def archivar_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id)
    expense.archivada = True
    expense.save()
    return redirect('expense')

def lista_expense(request):
    #empresa_id = request.GET.get('empresa')

    #empresa = get_object_or_404(Empresa, id=empresa_id)
    expenses = Expense.objects.all()
    
    totales_por_cotizacion = {}

    for expense in expenses:
        cotizacion = expense.voucher_number
        if cotizacion not in totales_por_cotizacion:
            totales_por_cotizacion[cotizacion] = {
                'subtotal': 0.0,
                'iva': 0.0,
                'total': 0.0,
            }
        totales_por_cotizacion[cotizacion]['subtotal'] += expense.subtotal or 0.0
        totales_por_cotizacion[cotizacion]['iva'] += expense.iva or 0.0
        totales_por_cotizacion[cotizacion]['total'] += expense.total or 0.0

    context = {
        'totales_por_cotizacion': totales_por_cotizacion,
        #'empresa': empresa,
    }
    return render(request, 'expense/lista_expense.html', context)

def detalle_expense(request, cotizacion):
    expenses = Expense.objects.filter(voucher_number=cotizacion)
    context = {
        'expenses': expenses,
        'cotizacion': cotizacion,
    }
    return render(request, 'expense/detalle.html', context)

def guardar_expense_client(request,): 
    if request.method == 'POST':
        formulario = ExpenseForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('guardar_expense_client')
    else:
        formulario = ExpenseForm()
    return render(request, 'expense/guardar_expense_client.html', {'formulario': formulario})

def listacli_expense(request, business_name):
    #empresa_id = request.GET.get('empresa')

    #empresa = get_object_or_404(Empresa, id=empresa_id)
    expenses = Expense.objects.filter(business_name)
    
    totales_por_client = {}

    for expense in expenses:
        business_name = expense.business_name
        if business_name not in totales_por_client:
            totales_por_client[business_name] = {
                'subtotal': 0.0,
                'iva': 0.0,
                'total': 0.0,
            }
        totales_por_client[business_name]['subtotal'] += expense.subtotal or 0.0
        totales_por_client[business_name]['iva'] += expense.iva or 0.0
        totales_por_client[business_name]['total'] += expense.total or 0.0

    context = {
        'totales_por_client': totales_por_client,
        #'empresa': empresa,
    }
    return render(request, 'expense/listacli_expense.html', context)

def detalle_client(request, clients):
    expenses = Expense.objects.filter(clients=clients)
    context = {
        'expenses': expenses,
        'client': clients,
    }
    return render(request, 'expense/detallecli.html', context)

def cotiza_expense(request, cotizacion):
    expenses = Expense.objects.filter(voucher_number=cotizacion).select_related('cotiza.business_name')
    context = {
        'expenses': expenses,
        'voucher_number': cotizacion,
        'client': client
    }
    return render(request, 'expense/cotizacion.html', context)

def exportar_pdf(request, cotizacion):
    cotizacion = Expense.objects.filter(voucher_number=cotizacion)
    empresa = Empresa.objects.first()
    context ={'cotizacion':cotizacion, 'empresa': empresa}
    html_string = render_to_string('cotiza/cotiza.html', context)
    html = HTML(string=html_string)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="cotiza_cotizacion_{}.pdf"'.format(cotizacion)
    html.write_pdf(response)
    return response

def exportar_excel_t(request):
    expenses = Expense.objects.all().values()
    df = pd.DataFrame(expenses)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cotizacion.xlsx"'
    df.to_excel(response, index=False)
    return response

def exportar_excel_c(request, cotizacion):
    expenses = Expense.objects.filter(voucher_number=cotizacion)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # Escribir el encabezado
    headers = ['Fecha', 'Cliente', 'Cotización', 'Cantidad', 'Descripción', 'Precio Unitario', 'Importe', 'no_instala', 'CI', 'Instalaciones', 'Subtotal', 'IVA', 'Total', 'Enviada', 'Aceptada', 'Facturada']
    ws.append(headers)

    # Escribir los datos
    for expense in expenses:
        cliente = expense.business_name
        producto = expense.product
        ws.append([
            expense.date,
            f"{cliente.business_name}",
            expense.voucher_number,
            expense.quantity,
            f"{producto.name}",
            expense.unit_price,
            expense.importe,
            expense.instalaciones,
            expense.cu_instala,
            expense.imp_instala,
            expense.subtotal,
            expense.iva,
            expense.total,
            expense.fecha_entrega,
            expense.fecha_autorizacion,
            expense.fecha_facturacion,
        ])

    # Responder con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=expenses_{cotizacion}.xlsx'
    wb.save(response)

    return response

def generate_excel(request, cotizacionf):
    expenses = Expense.objects.filter(voucher_number=cotizacionf)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Expenses'

    # Aplica formato previo a las celdas
    ws['C1'] = 'MARTHA PATRICIA MARTINEZ ALVARADO '
    ws['E1'] = 'COTIZACION NO.'
    ws['A6'] = 'FECHA'
    ws['A7'] = 'CLIENTE'
    ws['A8'] = 'CONTACTO'
    ws['E30'] = 'SUBTOTAL'
    ws['E31'] = 'IVA'
    ws['E32'] = 'TOTAL'
    ws['C30'] = 'PRECIOS VIGENTES HASTA FIN DE MES'
    ws['C31'] = '60% ANTICIPO 40% AL INSTALAR'
    ws['C32'] = 'ENTREGA 6 DÍAS HÁBILES'
    
    
    
    # Negrita
    font = Font(bold=True, size=14,)
    ws['C1'].font = font
    ws['E1'].font = font
    ws['A6'].font = font
    ws['A7'].font = font
    ws['A8'].font = font
    ws['E30'].font = font
    ws['E31'].font = font
    ws['E32'].font = font
    ws['C30'].font = font
    ws['C31'].font = font
    ws['C32'].font = font
    
    

    # Supongamos que 'empresa', 'fecha' y 'cliente' son atributos de un objeto Expense
    if expenses:
        expense = expenses[0]  # Obtén el primer objeto Expense (o cualquiera que prefieras)
        #ws['E1'] = expense.empresa  # Ajusta según el nombre del atributo de tu modelo
        cliente = expense.business_name
        ws['F1'] = expense.voucher_number
        ws['B6'] = expense.date.strftime('%Y-%m-%d')  # Formatea la fecha según tus necesidades
        ws['B7'] = f"{cliente.business_name}"
        ws['B8'] = f"{cliente.contact}"

    # Formato de celda (ejemplo: negrita)
    headers = ['CANTIDAD', 'UNIDAD', 'DESCRIPCIÓN', 'IMPORTE', 'INSTALACIONES', 'SUBTOTAL']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=10, column=col_num, value=header)

    # Escribir los datos a partir de la fila 11
    for row_num, expense in enumerate(expenses, 11):
        producto = expense.product
        measureunit = expense.measure_unit
        ws.cell(row=row_num, column=1, value=expense.quantity)
        ws.cell(row=row_num, column=2, value=f"{measureunit.description}")
        ws.cell(row=row_num, column=3, value=f"{producto.name}")
        ws.cell(row=row_num, column=4, value=expense.importe)
        ws.cell(row=row_num, column=5, value=expense.imp_instala)
        ws.cell(row=row_num, column=6, value=expense.subtotal)

    border = Border(left=Side(style='thick'),
                right=Side(style='thick'),
                top=Side(style='thick'),
                bottom=Side(style='thick'))
    ws['A10'].border = border
    ws['B10'].border = border
    ws['C10'].border = border
    ws['D10'].border = border
    ws['E10'].border = border
    ws['F10'].border = border 

    # Dar formato de número a una serie de celdas, de C11 a F32
    for row in ws.iter_rows(min_row=11, max_row=32, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'  # Formato de número con dos decimales

    # Dar un borde sencillo a una serie de celdas, por ejemplo, de A11 a H32
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=11, max_row=32, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
    
    # Color de fondo y borde grueso para la serie de celdas de A1 a F9
    fill = PatternFill(start_color="E0F8F7", end_color="E0F8F7", fill_type="solid")
    #thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    for row in ws.iter_rows(min_row=1, max_row=9, min_col=1, max_col=6):
        for cell in row:
            cell.fill = fill
            #cell.border = thick_border

    fill = PatternFill(start_color="E0F8F7", end_color="E0F8F7", fill_type="solid")
    #thick_border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

    for row in ws.iter_rows(min_row=30, max_row=32, min_col=3, max_col=3):
        for cell in row:
            cell.fill = fill
            

    # Inserta la imagen en la celda A1
    img = Image('libreria/templates/logo.png')
    ws.add_image(img, 'A1')

    # Inserta una fórmula en la celda H30, H31 y H32
    ws['F30'] = '=SUM(F3:F29)'
    ws['F31'] = '=(F30 * 0.16)'
    ws['F32'] = '=(F30 + F31)'

    # Ajustar el ancho de las columnas según el contenido
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtiene la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column].width = adjusted_width

    # Crea una respuesta HTTP con el contenido del archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=expenses_{cotizacionf}.xlsx'
    
    # Guarda el archivo Excel en la respuesta
    wb.save(response)

    return response

def crear_cotiza(request):
    if request.method == 'POST':
        form = CotizaForm(request.POST)
        if form.is_valid():
            form.save()
            # Redirige o muestra un mensaje de éxito
    else:
        form = CotizaForm()
    return render(request, 'cotiza/cotiza.html', {'form': form})

def ver_cotiza(request, cotiza_id):
    cotiza = Expense.objects.all() 
    return render(request, 'cotiza/cotiza.html', {'cotizas': cotizas})

def convert_cfdi_view(request):
    if request.method == 'POST':
        form = CFDIUploadForm(request.POST, request.FILES)
        if form.is_valid():
            xml_file = request.FILES['xml_file']
            cfdi_data = extract_cfdi_from_xml(xml_file)

            df = cfdi_to_excel(cfdi_data)

            # Generar respuesta con archivo Excel
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="cfdi.xlsx"'
            df.to_excel(response, index=False)

            return response
    else:
        form = CFDIUploadForm()

    return render(request, 'convert_cfdi.html', {'form': form})


        
def procesar_zip_view(request):
    if request.method == 'POST':
        form = CFDIZipForm(request.POST, request.FILES)
        if form.is_valid():
            zip_file = request.FILES['zip_file']
            zip_bytes = BytesIO(zip_file.read())

            all_rows = []
            with zipfile.ZipFile(zip_bytes, 'r') as z:
                for name in z.namelist():
                    if name.endswith('.xml'):
                        xml_content = z.read(name)
                        rows = extract_cfdi_from_xml(xml_content)
                        all_rows.extend(rows)

            df = pd.DataFrame(all_rows)

            # Generar nombre con timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cfdi_{timestamp}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'media/cfdis', filename)

            # Crear carpeta si no existe
            os.makedirs(os.path.dirname(file_path), exist_ok=True)

            # Guardar Excel en el servidor
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='CFDI Consolidado')

            return render(request, 'procesar_zip.html', {
                'form': CFDIZipForm(),
                'mensaje': f"✅ Archivo guardado exitosamente como: {filename}"
            })
    else:
        form = CFDIZipForm()
    
    return render(request, 'procesar_zip.html', {'form': form})

def historial_cfdis_view(request):
    carpeta = os.path.join(settings.MEDIA_ROOT, 'media/cfdis/')
    archivos = []

    if os.path.exists(carpeta):
        for nombre in sorted(os.listdir(carpeta), reverse=True):
            if nombre.endswith('.xlsx') and not nombre.startswith('~$'):
                ruta_archivo = os.path.join(carpeta, nombre)

                fecha_creacion = datetime.fromtimestamp(os.path.getctime(ruta_archivo))
                tam_kb = round(os.path.getsize(ruta_archivo) / 1024, 2)
                url = settings.MEDIA_URL + 'media/cfdis/' + nombre

                archivos.append({
                    'nombre': nombre,
                    'fecha': fecha_creacion.strftime('%Y-%m-%d %H:%M'),
                    'tamano': tam_kb,
                    'url': url
                })

    return render(request, 'cfdi/historial_cfdis.html', {'archivos': archivos})

def eliminar_cfdi_view(request, nombre):
    carpeta = os.path.join(settings.MEDIA_ROOT, 'media/cfdis/')
    ruta = os.path.join(carpeta, nombre)

    if os.path.exists(ruta) and ruta.endswith('.xlsx') and not nombre.startswith('~$'):
        os.remove(ruta)
        messages.success(request, f"Archivo '{nombre}' eliminado correctamente.")
    else:
        messages.error(request, f"No se pudo eliminar el archivo '{nombre}'.")

    return redirect('historial_cfdis')  # Asegúrate de que este nombre est